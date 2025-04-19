import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

def save_to_excel(immatriculation, date):
    """Saves immatriculation and date to an Excel file."""
    try:
        file_path = 'responses.xlsx' # Assuming the Excel file is in the root directory

        # Check if we have write permissions
        if os.path.exists(file_path):
            if not os.access(file_path, os.W_OK):
                print("Error: No write permissions for the Excel file")
                return
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            # Check if we can create new file
            try:
                # Attempt to create the file in the root directory
                with open(file_path, 'w') as test_file:
                    pass
                os.remove(file_path) # Remove the test file if creation was successful
            except PermissionError:
                print(f"Error: No permission to create Excel file at {file_path}")
                return
            except Exception as e:
                 print(f"Error during file creation check: {e}")
                 return

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(['Immatriculation', 'Date'])

        # Add data with RTL alignment for Arabic text
        row = [immatriculation, date]
        sheet.append(row)

        # Set RTL alignment for the immatriculation cell
        rtl_alignment = Alignment(horizontal='right', readingOrder=2)  # 2 = RTL
        sheet.cell(row=sheet.max_row, column=1).alignment = rtl_alignment

        workbook.save(file_path)

    except Exception as e:
        print(f"Error saving to Excel: {e}")
        print("Please ensure the file is not open in another program and you have write permissions")
